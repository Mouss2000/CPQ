import openpyxl
import re
import math

EXCEL_FILE = r'C:\Users\SetupGame\OneDrive\Desktop\prototype\PR ET COUT GRILLES LINEAIRE 20242025.xlsx'

def is_dimension(val):
    if not val: return False
    val_str = str(val).strip()
    return bool(re.match(r'^\d+[\*xX]\d+$', val_str))

def get_col_val(sheet, row, col_idx):
    cell = sheet.cell(row=row, column=col_idx)
    return cell.value

def process_table(sheet, start_col, name, ws_name):
    # start_col: 1 for A-E, 7 for G-K
    col_name = start_col
    col_qty = start_col + 2
    col_up = start_col + 3
    col_sub = start_col + 4
    
    blocks = []
    current_block = None
    
    for row_idx in range(1, sheet.max_row + 1):
        c_val = get_col_val(sheet, row_idx, col_name)
        if is_dimension(c_val):
            if current_block:
                blocks.append(current_block)
            current_block = {
                'sheet': ws_name,
                'dim': str(c_val),
                'components': [],
                'material_sum_excel': 0.0,
                'margin_pct': 0.0,
                'margin_row_found': False,
                'exec_time': 0.0,
                'exec_rate': 0.0,
                'exec_cost': 0.0,
                'total_pr': 0.0,
                'start_row': row_idx
            }
            continue
            
        if not current_block:
            continue
            
        c_val_str = str(c_val).strip().upper() if c_val else ""
        
        # Check components
        if c_val_str in ['PROFIL', 'AIL', 'T/ALUM', 'CLIP', 'EQUERE', 'REG', 'RENF', 'EMB', 'JOINT', 'COULISSE', 'LAME']:
            qty = get_col_val(sheet, row_idx, col_qty) or 0.0
            up = get_col_val(sheet, row_idx, col_up) or 0.0
            sub = get_col_val(sheet, row_idx, col_sub) or 0.0
            unit = str(get_col_val(sheet, row_idx, start_col + 1) or "").strip()
            
            try: qty = float(qty)
            except: qty = 0.0
            try: up = float(up)
            except: up = 0.0
            try: sub = float(sub)
            except: sub = 0.0
            
            current_block['components'].append({
                'name': c_val_str,
                'unit': unit,
                'qty': qty,
                'up': up,
                'sub': sub,
                'match': math.isclose(sub, qty * up, abs_tol=0.01)
            })
            continue
            
        # Check margin
        if 'FOUR' in c_val_str and '%' in c_val_str:
            current_block['margin_row_found'] = True
            m = re.search(r'(\d+(?:\.\d+)?)%', c_val_str)
            if m:
                current_block['margin_pct'] = float(m.group(1)) / 100.0
            continue
            
        # Check exec
        if 'EXEC' in c_val_str or 'MAIN D' in c_val_str:
            time = get_col_val(sheet, row_idx, col_qty) or 0.0
            rate = get_col_val(sheet, row_idx, col_up) or 0.0
            cost = get_col_val(sheet, row_idx, col_sub) or 0.0
            try: time = float(time)
            except: time = 0.0
            try: rate = float(rate)
            except: rate = 0.0
            try: cost = float(cost)
            except: cost = 0.0
            current_block['exec_time'] = time
            current_block['exec_rate'] = rate
            current_block['exec_cost'] = cost
            continue
            
        # Check total PR
        if 'PR' in c_val_str and ('TOTAL' in c_val_str or 'GRILLE' in c_val_str or len(c_val_str) < 5):
            val = get_col_val(sheet, row_idx, col_sub) or get_col_val(sheet, row_idx, col_sub - 1)
            try: total = float(val)
            except: total = 0.0
            if total > 0:
                current_block['total_pr'] = total
                blocks.append(current_block)
                current_block = None
            continue
            
    if current_block:
        blocks.append(current_block)
        
    return blocks

def main():
    print(f"LOADING EXCEL: {EXCEL_FILE}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        print(f"ERROR: {e}")
        return
        
    print(f"SHEETS: {wb.sheetnames}")
    
    all_blocks = []
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"--- SCANNING SHEET: {ws_name} ---")
        
        if 'FOUR' in ws_name.upper():
            print("FOUND FOUR MISE A JOUR SHEET")
            for row in range(1, ws.max_row + 1):
                mat = ws.cell(row=row, column=1).value
                price = ws.cell(row=row, column=2).value
                if mat and price:
                    print(f"MATERIAL: {mat} | PRICE: {price}")
            continue
            
        blocks_left = process_table(ws, 1, 'LEFT', ws_name)
        blocks_right = process_table(ws, 7, 'RIGHT', ws_name)
        
        all_blocks.extend(blocks_left)
        all_blocks.extend(blocks_right)
        
    print("\n================ BLOCK DETAILS ================\n")
    
    matched = 0
    mismatched = []
    margins = {}
    rates = {}
    comps = set()
    
    for b in all_blocks:
        mat_sum = sum(c['sub'] for c in b['components'])
        margined = mat_sum * (1 + b['margin_pct'])
        computed_pr = margined + b['exec_cost']
        
        pr_match = math.isclose(computed_pr, b['total_pr'], abs_tol=0.01)
        if pr_match:
            matched += 1
        else:
            mismatched.append(b)
            
        print(f"SHEET: {b['sheet']} | DIM: {b['dim']}")
        for c in b['components']:
            print(f"  COMP: {c['name']} | UNIT: {c['unit']} | QTY: {c['qty']} | UP: {c['up']} | SUB: {c['sub']} | MATH_OK: {c['match']}")
            comps.add(c['name'])
            
        print(f"  MAT_SUM: {mat_sum:.2f}")
        print(f"  MARGIN: {b['margin_pct']*100:.1f}% (Found: {b['margin_row_found']})")
        print(f"  MARGINED_MAT: {margined:.2f}")
        print(f"  EXEC: TIME {b['exec_time']} * RATE {b['exec_rate']} = {b['exec_cost']}")
        print(f"  COMPUTED_PR: {computed_pr:.2f} | EXCEL_PR: {b['total_pr']:.2f} | MATCH: {pr_match}\n")
        
        margins[b['margin_pct']] = margins.get(b['margin_pct'], 0) + 1
        rates[b['exec_rate']] = rates.get(b['exec_rate'], 0) + 1
        
    print("\n================ SUMMARY ================\n")
    print(f"TOTAL BLOCKS: {len(all_blocks)}")
    print(f"MATCHED PR: {matched}")
    print(f"MISMATCHED PR: {len(mismatched)}")
    
    print("\nMISMATCH DETAILS:")
    for b in mismatched:
        print(f"  {b['sheet']} {b['dim']} - Computed: {sum(c['sub'] for c in b['components']) * (1+b['margin_pct']) + b['exec_cost']:.2f}, Excel: {b['total_pr']:.2f}")
        
    print(f"\nMARGINS: {margins}")
    print(f"RATES: {rates}")
    print(f"COMPONENTS: {comps}")
    
if __name__ == "__main__":
    main()
