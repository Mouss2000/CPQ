"""
EXEC Rate Deep-Dive
Checks every EXEC row: is col_D always 200?
If not, is the sheet layout shifted (wrong column = wrong "rate")?
Groups by sheet to find patterns.
"""

import openpyxl
import os
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE_DIR, "PR ET COUT GRILLES LINEAIRE 20242025.xlsx")
wb = openpyxl.load_workbook(EXCEL, data_only=True)

dim_pattern = re.compile(r'^(\d+)\s*[\*xX]\s*(\d+)$')

# Collect every EXEC row with full context
exec_rows = []

for sname in wb.sheetnames:
    ws = wb[sname]
    for r in range(1, ws.max_row + 1):
        for base_col in [1, 7]:  # left table (col A) and right table (col G)
            cell_b = ws.cell(row=r, column=base_col + 1).value
            if cell_b and isinstance(cell_b, str) and 'EXEC' in cell_b.upper():
                # Grab cols A-F (or G-L) for full context
                context = {}
                for offset in range(6):
                    col = base_col + offset
                    val = ws.cell(row=r, column=col).value
                    col_letter = chr(ord('A') + offset)  # relative: A,B,C,D,E,F
                    context[col_letter] = val
                
                # Also grab 2 rows above for dimension context
                dim_text = None
                for look_back in range(1, 10):
                    prev_val = ws.cell(row=r - look_back, column=base_col).value
                    if prev_val and isinstance(prev_val, str) and dim_pattern.match(prev_val.strip()):
                        dim_text = prev_val.strip()
                        break
                
                rate = context.get('D')
                hours = context.get('C')
                total = context.get('E')
                
                # Verify: does C * D = E?
                calc_ok = None
                if isinstance(hours, (int, float)) and isinstance(rate, (int, float)) and isinstance(total, (int, float)):
                    expected = hours * rate
                    calc_ok = abs(expected - total) < 0.01
                
                exec_rows.append({
                    'sheet': sname,
                    'row': r,
                    'base_col': base_col,
                    'dim': dim_text,
                    'hours': hours,
                    'rate': rate,
                    'total': total,
                    'cxd_eq_e': calc_ok,
                    'context': context,
                })

# ---- ANALYSIS ----
print(f"Total EXEC rows found: {len(exec_rows)}")
print()

# Group by rate value
from collections import defaultdict
by_rate = defaultdict(list)
for ex in exec_rows:
    r = ex['rate']
    if isinstance(r, (int, float)):
        by_rate[round(r, 2)].append(ex)
    else:
        by_rate[f"non-numeric:{repr(r)}"].append(ex)

print("=== EXEC RATE DISTRIBUTION (with sheet breakdown) ===")
for rate_key in sorted(by_rate.keys(), key=lambda x: (isinstance(x, str), x)):
    items = by_rate[rate_key]
    sheets = defaultdict(int)
    for it in items:
        sheets[it['sheet']] += 1
    sheet_str = ", ".join(f"{s}({c})" for s, c in sorted(sheets.items(), key=lambda x: -x[1]))
    
    # Check C*D=E consistency
    ok_count = sum(1 for it in items if it['cxd_eq_e'] is True)
    fail_count = sum(1 for it in items if it['cxd_eq_e'] is False)
    
    print(f"\n  Rate={rate_key}: {len(items)} products | C*D=E: {ok_count} ok, {fail_count} fail")
    print(f"    Sheets: {sheet_str}")
    
    # Show 3 examples
    for ex in items[:3]:
        print(f"    Ex: [{ex['sheet']}] row {ex['row']} dim={ex['dim']} | C(hrs)={ex['hours']} D(rate)={ex['rate']} E(total)={ex['total']} | C*D=E: {ex['cxd_eq_e']}")

# ---- Focus on non-200 rates: are they column misreads? ----
print("\n\n=== NON-200 RATE DETAILED INSPECTION ===")
print("Showing full row context for non-200 rates to check column alignment\n")

non200 = [ex for ex in exec_rows if isinstance(ex['rate'], (int, float)) and abs(ex['rate'] - 200) > 1]

# Group by sheet
non200_sheets = defaultdict(list)
for ex in non200:
    non200_sheets[ex['sheet']].append(ex)

for sheet_name, items in sorted(non200_sheets.items()):
    print(f"\n  Sheet: {sheet_name} ({len(items)} non-200 EXEC rows)")
    for ex in items[:5]:  # first 5 per sheet
        ctx = ex['context']
        print(f"    Row {ex['row']} (base_col={ex['base_col']}) dim={ex['dim']}:")
        print(f"      A={repr(ctx['A'])} B={repr(ctx['B'])} C={repr(ctx['C'])} D={repr(ctx['D'])} E={repr(ctx['E'])} F={repr(ctx.get('F'))}")
        
        # Check: maybe the REAL rate is in a different column?
        # Try col E or col F as rate, or check if D is actually hours
        if isinstance(ctx['D'], (int, float)) and isinstance(ctx['C'], (int, float)):
            print(f"      C*D = {ctx['C'] * ctx['D']:.2f} vs E = {ctx['E']}")
            if isinstance(ctx['E'], (int, float)) and ctx['C'] != 0:
                print(f"      E/C = {ctx['E'] / ctx['C']:.2f} (implied rate if E=hours*rate)")

print("\n\n=== RATE=200 VERIFICATION (spot check) ===")
rate200 = [ex for ex in exec_rows if isinstance(ex['rate'], (int, float)) and abs(ex['rate'] - 200) < 1]
print(f"Total with rate=200: {len(rate200)}")
# Verify C*200=E for all
mismatches = [ex for ex in rate200 if ex['cxd_eq_e'] is False]
print(f"C*200=E matches: {len(rate200) - len(mismatches)}")
print(f"C*200=E fails: {len(mismatches)}")

# Show hours distribution for rate=200 products
hours_dist = defaultdict(int)
for ex in rate200:
    if isinstance(ex['hours'], (int, float)):
        hours_dist[ex['hours']] += 1

print(f"\nExecution time distribution (rate=200 products):")
for h in sorted(hours_dist.keys()):
    mins = h * 60
    print(f"  {h}h ({mins:.0f}min): {hours_dist[h]} products")
