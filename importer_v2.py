"""
CPQ Importer V2 — Fixed extraction logic.
Correctly reads:
  1. TOTAL rows (full BOM sum) instead of PROFIL-only rows
  2. Side-by-side tables (Blanc cols A-E, Gris cols G-K)
  3. HSD sheets (single-table, separate per color)
  4. Tariff sheet
"""

import openpyxl
import sqlite3
import os
import re
from database_manager import DatabaseManager


class CPQImporterV2:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.db_mgr = DatabaseManager()
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)

    def run(self):
        print(f"[V2] Starting import from {self.excel_path}...")
        conn = self.db_mgr.get_connection()
        cursor = conn.cursor()

        # Wipe existing product data for clean re-import
        cursor.execute("DELETE FROM products")
        cursor.execute('''CREATE TABLE IF NOT EXISTS product_components (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            component_name TEXT NOT NULL,
            quantity REAL,
            unit_price REAL,
            subtotal REAL,
            FOREIGN KEY (product_id) REFERENCES products(id)
        )''')
        cursor.execute('DELETE FROM product_components')
        cursor.execute("DELETE FROM product_categories")
        conn.commit()

        self._seed_categories(cursor)
        conn.commit()

        # Process each sheet
        for sheet_name in self.wb.sheetnames:
            if any(p in sheet_name for p in ['GSD', 'GDD', 'GBF']):
                if 'TARIF' in sheet_name.upper():
                    continue
                self._process_sheet(cursor, sheet_name)

        conn.commit()

        # Update tariffs
        self._update_tariffs(cursor)
        conn.commit()

        # Report
        cursor.execute("SELECT COUNT(*) FROM products")
        total = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM products WHERE excel_tariff IS NOT NULL")
        with_tariff = cursor.fetchone()[0]
        print(f"\n[V2] Import complete: {total} products, {with_tariff} with tariffs")

        # Validation sample
        print("\n[V2] Sample verification:")
        cursor.execute("""
            SELECT category_code, color, dimension, excel_cost, excel_tariff 
            FROM products 
            WHERE dimension = '600*60' AND category_code = 'GBF'
        """)
        for r in cursor.fetchall():
            print(f"  {r}")

        conn.close()

    def _seed_categories(self, cursor):
        categories = [
            ('GSD', 'Grilles Simple Deflexion'),
            ('GDD', 'Grilles Double Deflexion'),
            ('GBF', 'Grilles A Barre Fixe'),
        ]
        for code, name in categories:
            cursor.execute(
                "INSERT OR IGNORE INTO product_categories (code, name) VALUES (?, ?)",
                (code, name)
            )

    def _detect_category(self, sheet_name):
        """Detect product category from sheet name."""
        sn = sheet_name.upper()
        if 'GBF' in sn:
            return 'GBF'
        elif 'GDD' in sn:
            return 'GDD'
        elif 'GSD' in sn:
            return 'GSD'
        return None

    def _detect_sheet_layout(self, sheet_name, sheet):
        """Detect whether sheet has side-by-side tables and which colors."""
        sn = sheet_name.upper()
        max_col = sheet.max_column

        # Check if there's data in columns G+ (side-by-side)
        has_right_table = False
        for r in range(1, min(20, sheet.max_row + 1)):
            for c in range(7, min(max_col + 1, 13)):
                v = sheet.cell(row=r, column=c).value
                if v is not None and str(v).strip():
                    has_right_table = True
                    break
            if has_right_table:
                break

        # Determine color assignments
        # Look at sheet name and row 3-4 headers for clues
        left_color = None
        right_color = None

        if has_right_table:
            # Check headers for color info
            for r in range(1, 6):
                for c in range(1, max_col + 1):
                    v = sheet.cell(row=r, column=c).value
                    if v and isinstance(v, str):
                        vu = v.upper()
                        if c <= 6:
                            if 'BLANC' in vu or ' BL' in vu:
                                left_color = 'Blanc'
                            elif 'GRIS' in vu or ' GR' in vu:
                                left_color = 'Gris Anodise'
                        else:
                            if 'BLANC' in vu or ' BL' in vu:
                                right_color = 'Blanc'
                            elif 'GRIS' in vu or ' GR' in vu:
                                right_color = 'Gris Anodise'
                            elif 'DOUBLE' in vu:
                                right_color = '__GDD__'  # Different category, not color

            # Fallback from sheet name
            if 'GR ET BL' in sn or 'GR  ET BL' in sn or ('GRIS' in sn and 'BLANC' in sn):
                left_color = left_color or 'Blanc'
                right_color = right_color or 'Gris Anodise'
            elif 'SIMPLE' in sn and 'DOUBLE' in sn:
                left_color = left_color or 'Blanc'
                right_color = right_color or 'Blanc'
        else:
            # Single table — determine color from sheet name
            if 'BL' in sn or 'BLANC' in sn:
                left_color = 'Blanc'
            elif 'GR' in sn or 'GRIS' in sn:
                left_color = 'Gris Anodise'
            else:
                left_color = 'Blanc'  # Default

        return has_right_table, left_color, right_color

    def _parse_dimension(self, dim_str):
        """Extract width*height from dimension string."""
        if not dim_str:
            return None, None
        match = re.search(r'(\d+)\s*\*\s*(\d+)', str(dim_str))
        if match:
            return int(match.group(1)), int(match.group(2))
        return None, None

    def _extract_blocks_from_table(self, sheet, dim_col, total_col, start_row=1):
        """
        Extract (dimension_str, w, h, total_cost, components) tuples from a single table column range.
        
        Strategy: Find dimension rows (have W*H pattern in dim_col), then find 
        the TOTAL row = last row with a numeric value in total_col before the next
        dimension row or blank section.
        
        components is a list of (name, quantity, unit_price, subtotal) tuples.
        Quantity is stored as a float when numeric, or None for non-numeric values like '15 min'.
        """
        blocks = []
        max_row = sheet.max_row

        # First pass: find all dimension row positions
        dim_rows = []
        for r in range(start_row, max_row + 1):
            v = sheet.cell(row=r, column=dim_col).value
            if v:
                w, h = self._parse_dimension(v)
                if w and h:
                    dim_rows.append((r, f"{w}*{h}", w, h))

        # Second pass: for each dimension, find the TOTAL row and collect components
        for i, (dim_row, dim_str, w, h) in enumerate(dim_rows):
            # Search from dim_row+1 to next dim_row (or +15 rows max)
            if i + 1 < len(dim_rows):
                end_search = dim_rows[i + 1][0]
            else:
                end_search = min(dim_row + 15, max_row + 1)

            total_cost = None
            components = []
            for r in range(dim_row, end_search):
                cell_a = sheet.cell(row=r, column=dim_col).value
                cell_b = sheet.cell(row=r, column=dim_col + 1).value
                cell_e = sheet.cell(row=r, column=total_col).value

                # TOTAL row pattern: dim_col and dim_col+1 are None, total_col has number
                if cell_a is None and cell_b is None and isinstance(cell_e, (int, float)):
                    total_cost = cell_e  # Keep updating — last one is the final total (after EXEC)
                elif cell_b is not None and isinstance(cell_b, str) and cell_b.strip():
                    # Component row: read name, quantity, unit_price, subtotal
                    comp_name = cell_b.strip()
                    raw_qty = sheet.cell(row=r, column=dim_col + 2).value
                    comp_unit_price = sheet.cell(row=r, column=dim_col + 3).value
                    comp_subtotal = cell_e  # total_col = dim_col + 4

                    # Only store numeric quantities; non-numeric like '15 min' -> None
                    if isinstance(raw_qty, (int, float)):
                        comp_qty = float(raw_qty)
                    else:
                        comp_qty = None

                    components.append((comp_name, comp_qty, comp_unit_price, comp_subtotal))

            if total_cost is not None:
                blocks.append((dim_str, w, h, total_cost, components))
            else:
                # Fallback: no total row found, skip
                pass

        return blocks

    def _detect_categories_from_headers(self, sheet, default_category):
        """Scan header rows to detect left/right table categories."""
        left_cat = default_category
        right_cat = default_category
        for r in range(1, 6):
            for c in range(1, min(sheet.max_column + 1, 18)):
                v = sheet.cell(row=r, column=c).value
                if v and isinstance(v, str):
                    vu = v.upper()
                    if 'SIMPLE' in vu and 'DEFLEXION' in vu:
                        if c <= 6:
                            left_cat = 'GSD'
                        elif c <= 12:  # Only cols 7-12 for right table
                            right_cat = 'GSD'
                    elif 'DOUBLE' in vu and 'DEFLEXION' in vu:
                        if c <= 6:
                            left_cat = 'GDD'
                        elif c <= 12:
                            right_cat = 'GDD'
                    elif 'BARRE FIXE' in vu:
                        if c <= 6:
                            left_cat = 'GBF'
                        elif c <= 12:
                            right_cat = 'GBF'
        return left_cat, right_cat

    def _process_sheet(self, cursor, sheet_name):
        sheet = self.wb[sheet_name]
        category = self._detect_category(sheet_name)
        if not category:
            return

        has_right, left_color, right_color = self._detect_sheet_layout(sheet_name, sheet)
        left_cat, right_cat = self._detect_categories_from_headers(sheet, category)

        print(f"  Processing: {sheet_name}")
        print(f"    Layout={'side-by-side' if has_right else 'single'}")
        print(f"    Left: {left_cat} {left_color}, Right: {right_cat} {right_color}")

        # Left table: cols A=1, E=5
        left_blocks = self._extract_blocks_from_table(sheet, dim_col=1, total_col=5)
        count_left = 0
        for dim_str, w, h, cost, components in left_blocks:
            cursor.execute("""
                INSERT INTO products (category_code, dimension, width, height, color, excel_cost)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (left_cat, dim_str, w, h, left_color, cost))
            product_id = cursor.lastrowid
            for comp_name, comp_qty, comp_unit_price, comp_subtotal in components:
                cursor.execute("""
                    INSERT INTO product_components (product_id, component_name, quantity, unit_price, subtotal)
                    VALUES (?, ?, ?, ?, ?)
                """, (product_id, comp_name, comp_qty, comp_unit_price, comp_subtotal))
            count_left += 1

        # Right table (if exists)
        count_right = 0
        if has_right:
            # Detect right table column offset — usually G=7, K=11
            # But some sheets use H=8, L=12
            right_dim_col = None
            for test_col in [7, 8]:
                for r in range(1, min(30, sheet.max_row + 1)):
                    v = sheet.cell(row=r, column=test_col).value
                    if v:
                        wt, ht = self._parse_dimension(v)
                        if wt and ht:
                            right_dim_col = test_col
                            break
                if right_dim_col:
                    break

            if right_dim_col:
                right_total_col = right_dim_col + 4  # 4 cols to the right
                r_color = right_color if right_color and right_color != '__GDD__' else left_color

                right_blocks = self._extract_blocks_from_table(
                    sheet, dim_col=right_dim_col, total_col=right_total_col
                )
                for dim_str, w, h, cost, components in right_blocks:
                    cursor.execute("""
                        INSERT INTO products (category_code, dimension, width, height, color, excel_cost)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (right_cat, dim_str, w, h, r_color, cost))
                    product_id = cursor.lastrowid
                    for comp_name, comp_qty, comp_unit_price, comp_subtotal in components:
                        cursor.execute("""
                            INSERT INTO product_components (product_id, component_name, quantity, unit_price, subtotal)
                            VALUES (?, ?, ?, ?, ?)
                        """, (product_id, comp_name, comp_qty, comp_unit_price, comp_subtotal))
                    count_right += 1

        print(f"    Extracted: {count_left} left, {count_right} right")

    def _update_tariffs(self, cursor):
        """Update tariffs from the TARIF sheet."""
        print("  Updating tariffs from 'TARIF GRILLES ALI BABA'...")
        if 'TARIF GRILLES ALI BABA' not in self.wb.sheetnames:
            print("    Tariff sheet not found, skipping.")
            return

        sheet = self.wb['TARIF GRILLES ALI BABA']
        updated = 0
        for row in sheet.iter_rows(min_row=5, max_col=3, values_only=True):
            dim_str, price_bl, price_gr = row
            if dim_str:
                width, height = self._parse_dimension(dim_str)
                if width and height:
                    if price_bl is not None:
                        cursor.execute("""
                            UPDATE products SET excel_tariff = ?
                            WHERE width = ? AND height = ? AND color = 'Blanc'
                        """, (price_bl, width, height))
                        updated += cursor.rowcount

                    if price_gr is not None:
                        cursor.execute("""
                            UPDATE products SET excel_tariff = ?
                            WHERE width = ? AND height = ? AND color = 'Gris Anodise'
                        """, (price_gr, width, height))
                        updated += cursor.rowcount

        print(f"    Tariffs updated: {updated} rows")


if __name__ == "__main__":
    importer = CPQImporterV2('PR ET COUT GRILLES LINEAIRE 20242025.xlsx')
    importer.run()
