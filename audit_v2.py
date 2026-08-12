"""
audit_v2.py — Deep Excel formula audit.
Phase 1: Raw cell dump around dimension rows to map actual layout.
Phase 2: Full formula verification with correct column mapping.
"""
import openpyxl
import re
import math
from collections import defaultdict

EXCEL = r'C:\Users\SetupGame\OneDrive\Desktop\prototype\PR ET COUT GRILLES LINEAIRE 20242025.xlsx'

def is_dim(v):
    if not v: return False
    return bool(re.match(r'^\d+[\*xX]\d+$', str(v).strip()))

def safe_float(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def dump_raw_blocks(wb, max_blocks=5):
    """Phase 1: Dump raw cells around first few dimension rows per sheet to understand layout."""
    print("="*80)
    print("PHASE 1: RAW CELL DUMP (first 5 blocks per sheet)")
    print("="*80)
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"\n{'='*60}")
        print(f"SHEET: {ws_name}  (rows={ws.max_row}, cols={ws.max_column})")
        print(f"{'='*60}")
        
        # Dump header rows (1-8) to see column structure
        print("HEADER ROWS (1-8):")
        for r in range(1, min(9, ws.max_row+1)):
            vals = []
            for c in range(1, min(ws.max_column+1, 13)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    vals.append(f"[{c}]={v}")
            if vals:
                print(f"  R{r}: {' | '.join(vals)}")
        
        # Find dimension rows and dump surrounding context
        block_count = 0
        for r in range(1, ws.max_row+1):
            # Check cols A and G for dimensions
            for start_col in [1, 7]:
                v = ws.cell(row=r, column=start_col).value
                if is_dim(v) and block_count < max_blocks:
                    block_count += 1
                    print(f"\n  BLOCK #{block_count} at row {r}, col_offset={start_col}, dim={v}")
                    # Dump 12 rows starting from dim row
                    for dr in range(r, min(r+14, ws.max_row+1)):
                        vals = []
                        for c in range(start_col, min(start_col+5, ws.max_column+1)):
                            cv = ws.cell(row=dr, column=c).value
                            if cv is not None:
                                vals.append(f"[c{c}]={cv}")
                        if vals:
                            print(f"    R{dr}: {' | '.join(vals)}")
            if block_count >= max_blocks:
                break

def full_audit(wb):
    """Phase 2: Full formula verification."""
    print("\n" + "="*80)
    print("PHASE 2: FULL FORMULA AUDIT")
    print("="*80)
    
    all_blocks = []
    all_comp_names = set()
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        # Skip FOUR MISE A JOUR — it's a reference table, not product blocks
        if 'FOUR' in ws_name.upper():
            print(f"\n--- FOUR MISE A JOUR SHEET: {ws_name} ---")
            for r in range(1, ws.max_row+1):
                row_vals = []
                for c in range(1, min(ws.max_column+1, 8)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        row_vals.append(f"[c{c}]={v}")
                if row_vals:
                    print(f"  R{r}: {' | '.join(row_vals)}")
            continue
        
        for start_col in [1, 7]:
            blocks = extract_blocks(ws, ws_name, start_col, all_comp_names)
            all_blocks.extend(blocks)
    
    # Verification
    matched = 0
    mismatched = []
    margins = defaultdict(int)
    rates = defaultdict(int)
    exec_match_count = 0
    exec_total = 0
    comp_math_ok = 0
    comp_math_total = 0
    
    for b in all_blocks:
        mat_sum = sum(c['sub'] for c in b['components'])
        margined = mat_sum * (1 + b['margin_pct'])
        exec_computed = b['exec_time'] * b['exec_rate']
        computed_pr = margined + b['exec_cost']
        
        # Check exec math
        if b['exec_time'] > 0:
            exec_total += 1
            if math.isclose(exec_computed, b['exec_cost'], abs_tol=0.02):
                exec_match_count += 1
        
        # Check component math
        for c in b['components']:
            comp_math_total += 1
            if c['match']:
                comp_math_ok += 1
        
        # Check PR
        if b['total_pr'] > 0:
            pr_match = math.isclose(computed_pr, b['total_pr'], abs_tol=0.5)
            if pr_match:
                matched += 1
            else:
                mismatched.append({
                    'sheet': b['sheet'], 'dim': b['dim'],
                    'mat_sum': mat_sum, 'margin': b['margin_pct'],
                    'margined': margined, 'exec_cost': b['exec_cost'],
                    'computed': computed_pr, 'excel': b['total_pr'],
                    'diff': abs(computed_pr - b['total_pr']),
                    'components': b['components']
                })
        
        margins[b['margin_pct']] += 1
        rates[b['exec_rate']] += 1
    
    # Summary
    total_with_pr = sum(1 for b in all_blocks if b['total_pr'] > 0)
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"Total blocks found: {len(all_blocks)}")
    print(f"Blocks with PR > 0: {total_with_pr}")
    print(f"PR MATCHED: {matched}/{total_with_pr} ({100*matched/max(total_with_pr,1):.1f}%)")
    print(f"PR MISMATCHED: {len(mismatched)}")
    print(f"Component math (sub=qty*up): {comp_math_ok}/{comp_math_total} ({100*comp_math_ok/max(comp_math_total,1):.1f}%)")
    print(f"EXEC math (cost=time*rate): {exec_match_count}/{exec_total} ({100*exec_match_count/max(exec_total,1):.1f}%)")
    
    print(f"\nMARGIN DISTRIBUTION:")
    for m, cnt in sorted(margins.items()):
        print(f"  {m*100:.1f}%: {cnt} blocks")
    
    print(f"\nEXEC RATE DISTRIBUTION:")
    for r, cnt in sorted(rates.items()):
        if cnt >= 2:
            print(f"  {r} DH/h: {cnt} blocks")
    
    print(f"\nALL COMPONENT NAMES FOUND: {sorted(all_comp_names)}")
    
    if mismatched:
        print(f"\nFIRST 30 MISMATCHES (of {len(mismatched)}):")
        for m in mismatched[:30]:
            print(f"  {m['sheet']} | {m['dim']} | mat_sum={m['mat_sum']:.2f} | margin={m['margin']*100:.0f}% | margined={m['margined']:.2f} | exec={m['exec_cost']:.2f} | computed={m['computed']:.2f} | excel={m['excel']:.2f} | diff={m['diff']:.2f}")
            for c in m['components']:
                print(f"    {c['name']:10s} qty={c['qty']:8.3f} up={c['up']:8.2f} sub={c['sub']:8.2f} ok={c['match']}")
    
    return all_blocks, mismatched


def extract_blocks(ws, ws_name, start_col, all_comp_names):
    """Extract product blocks from a table starting at start_col."""
    blocks = []
    
    # Known component names (expanded list)
    COMP_NAMES = {'PROFIL', 'AIL', 'T/ALUM', 'CLIP', 'EQUERE', 'REG', 'RENF',
                  'EMB', 'JOINT', 'COULISSE', 'LAME', 'EXEC', 'MAIN D\'OEUVRE',
                  'RIVET', 'VIS', 'AXE', 'RESSORT', 'CHARNIERE', 'GOUPILLE',
                  'CADRE', 'TRAVERSE', 'MONTANT', 'TOLE'}
    
    # Find all dimension rows
    dim_rows = []
    for r in range(1, ws.max_row+1):
        v = ws.cell(row=r, column=start_col).value
        if is_dim(v):
            dim_rows.append(r)
    
    for i, dim_row in enumerate(dim_rows):
        dim_val = str(ws.cell(row=dim_row, column=start_col).value).strip()
        
        # Scan forward from dim_row to find components, margin, exec, total
        # Stop at next dimension row or after 20 rows
        end_row = dim_rows[i+1] if i+1 < len(dim_rows) else min(dim_row + 20, ws.max_row+1)
        
        block = {
            'sheet': ws_name, 'dim': dim_val,
            'components': [],
            'margin_pct': 0.0, 'margin_row_found': False,
            'exec_time': 0.0, 'exec_rate': 0.0, 'exec_cost': 0.0,
            'total_pr': 0.0
        }
        
        # The dimension row itself might have PROFIL data in col B (or start_col+1)
        for r in range(dim_row, end_row):
            # Read all 5 columns
            c_a = ws.cell(row=r, column=start_col).value      # col A/G: dim or empty
            c_b = ws.cell(row=r, column=start_col+1).value    # col B/H: component name
            c_c = ws.cell(row=r, column=start_col+2).value    # col C/I: qty
            c_d = ws.cell(row=r, column=start_col+3).value    # col D/J: unit price / rate
            c_e = ws.cell(row=r, column=start_col+4).value    # col E/K: subtotal
            
            # Normalize component name — could be in col A or col B
            comp_name = None
            for candidate in [c_b, c_a]:
                if candidate and str(candidate).strip():
                    cand_upper = str(candidate).strip().upper()
                    # Check known names or partial matches
                    for kn in COMP_NAMES:
                        if kn in cand_upper or cand_upper in kn:
                            comp_name = kn
                            break
                    if not comp_name and len(cand_upper) > 1 and cand_upper not in dim_val:
                        # Could be an unknown component
                        if c_c is not None and c_d is not None and c_e is not None:
                            comp_name = cand_upper
                    if comp_name:
                        break
            
            # Check for EXEC
            c_b_str = str(c_b).strip().upper() if c_b else ""
            c_a_str = str(c_a).strip().upper() if c_a else ""
            
            is_exec = ('EXEC' in c_b_str or 'EXEC' in c_a_str or 
                       'MAIN D' in c_b_str or 'MAIN D' in c_a_str)
            
            # Check for margin row ("four X%")
            for cell_val in [c_a, c_b, c_d]:
                if cell_val:
                    cell_str = str(cell_val).strip().upper()
                    if 'FOUR' in cell_str and '%' in cell_str:
                        block['margin_row_found'] = True
                        m = re.search(r'(\d+(?:\.\d+)?)\s*%', cell_str)
                        if m:
                            block['margin_pct'] = float(m.group(1)) / 100.0
                    # Also check "four" in col D with just a percentage
                    elif 'FOUR' in cell_str:
                        # Check if col D has percentage
                        d_str = str(c_d).strip() if c_d else ""
                        m2 = re.search(r'(\d+(?:\.\d+)?)\s*%', d_str)
                        if m2:
                            block['margin_row_found'] = True
                            block['margin_pct'] = float(m2.group(1)) / 100.0
            
            # Also check col D alone for "four X%"
            if c_d:
                d_str = str(c_d).strip().upper()
                if 'FOUR' in d_str:
                    m3 = re.search(r'(\d+(?:\.\d+)?)\s*%', d_str)
                    if m3:
                        block['margin_row_found'] = True
                        block['margin_pct'] = float(m3.group(1)) / 100.0
            
            if is_exec:
                block['exec_time'] = safe_float(c_c)
                block['exec_rate'] = safe_float(c_d)
                block['exec_cost'] = safe_float(c_e)
                continue
            
            if comp_name and comp_name != 'EXEC':
                qty = safe_float(c_c)
                up = safe_float(c_d)
                sub = safe_float(c_e)
                
                # Only add if at least one numeric value exists
                if qty > 0 or up > 0 or sub > 0:
                    all_comp_names.add(comp_name)
                    match = math.isclose(sub, qty * up, abs_tol=0.02) if qty > 0 and up > 0 else True
                    block['components'].append({
                        'name': comp_name, 'qty': qty, 'up': up, 'sub': sub, 'match': match
                    })
                continue
            
            # Check for total PR row — look for numeric value in col E with no component
            # Usually the last row before next dimension, with col B/A empty or containing "PR"/"TOTAL"
            if c_e is not None and comp_name is None and not is_exec:
                for check in [c_a_str, c_b_str]:
                    if 'PR' in check or 'TOTAL' in check or 'GRILLE' in check:
                        val = safe_float(c_e)
                        if val > 0:
                            block['total_pr'] = val
                            break
                # Also: if this is a row with ONLY col E having a value (summary row)
                if block['total_pr'] == 0 and c_a is None and c_b is None and c_c is None and c_d is None:
                    val = safe_float(c_e)
                    if val > 0:
                        # Could be material sum or total — check position
                        pass  # Skip ambiguous solo values
        
        # Only keep blocks with actual data
        if block['components'] or block['total_pr'] > 0:
            blocks.append(block)
    
    return blocks


def main():
    print(f"Loading: {EXCEL}")
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")
    
    dump_raw_blocks(wb, max_blocks=3)
    all_blocks, mismatched = full_audit(wb)

if __name__ == "__main__":
    main()
